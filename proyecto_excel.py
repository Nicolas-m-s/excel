import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import re

# Archivo Excel
nombre_archivo = "datos.xlsx"

# Crear o cargar archivo
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"])
    wb.save(nombre_archivo)


def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    # Validar campos vacíos
    if not all([nombre, edad, email, telefono, direccion]):
        messagebox.showwarning(
            title="Advertencia",
            message="Todos los campos son obligatorios"
        )
        return

    # Validar números
    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(
            title="Advertencia",
            message="Edad y teléfono deben ser números"
        )
        return

    # Validar correo
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning(
            title="Advertencia",
            message="Correo electrónico no válido"
        )
        return

    # Guardar en Excel
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(nombre_archivo)

    messagebox.showinfo(
        title="Éxito",
        message="Datos guardados correctamente"
    )

    # Limpiar campos
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)


# Crear ventana
ventana = tk.Tk()
ventana.title("Formulario de Datos")
ventana.geometry("900x700")
ventana.configure(bg="light green")

# Configuración de columnas
ventana.columnconfigure(0, weight=1)
ventana.columnconfigure(1, weight=3)

# Estilos
fuente_grande = ("Times New Roman", 14)
color_fondo = "light green"

# Campos
tk.Label(
    ventana,
    text="Nombre:",
    bg=color_fondo,
    fg="black",
    font=fuente_grande
).grid(row=0, column=0, padx=10, pady=10, sticky="w")

entry_nombre = tk.Entry(ventana, font=fuente_grande)
entry_nombre.grid(row=0, column=1, padx=10, pady=10, sticky="ew")


tk.Label(
    ventana,
    text="Edad:",
    bg=color_fondo,
    fg="black",
    font=fuente_grande
).grid(row=1, column=0, padx=10, pady=10, sticky="w")

entry_edad = tk.Entry(ventana, font=fuente_grande)
entry_edad.grid(row=1, column=1, padx=10, pady=10, sticky="ew")


tk.Label(
    ventana,
    text="Email:",
    bg=color_fondo,
    fg="black",
    font=fuente_grande
).grid(row=2, column=0, padx=10, pady=10, sticky="w")

entry_email = tk.Entry(ventana, font=fuente_grande)
entry_email.grid(row=2, column=1, padx=10, pady=10, sticky="ew")


tk.Label(
    ventana,
    text="Teléfono:",
    bg=color_fondo,
    fg="black",
    font=fuente_grande
).grid(row=3, column=0, padx=10, pady=10, sticky="w")

entry_telefono = tk.Entry(ventana, font=fuente_grande)
entry_telefono.grid(row=3, column=1, padx=10, pady=10, sticky="ew")


tk.Label(
    ventana,
    text="Dirección:",
    bg=color_fondo,
    fg="black",
    font=fuente_grande
).grid(row=4, column=0, padx=10, pady=10, sticky="w")

entry_direccion = tk.Entry(ventana, font=fuente_grande)
entry_direccion.grid(row=4, column=1, padx=10, pady=10, sticky="ew")


# Botón guardar
btn_guardar = tk.Button(
    ventana,
    text="Guardar",
    command=guardar_datos,
    bg="white",
    fg="black",
    font=("Times New Roman", 14, "bold")
)

btn_guardar.grid(row=5, column=0, columnspan=2, pady=30)

ventana.mainloop()